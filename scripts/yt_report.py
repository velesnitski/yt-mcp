#!/usr/bin/env python3
"""
YouTrack Multi-Team — Daily Issue Report Generator.

Fetches unresolved + recently resolved issues from YouTrack for each team,
compares with the previous report, and generates an Excel file with delta
tracking. Optionally fetches daily GitLab commit statistics per team.

Environment variables:
    YOUTRACK_TOKEN  — YouTrack permanent token (required)
    YOUTRACK_URL    — YouTrack base URL (required)
    GITLAB_TOKEN    — GitLab private token (optional, for commit stats)
    GITLAB_URL      — GitLab base URL (optional, for commit stats)
    TEAMS_CONFIG    — Path to teams JSON config file (default: scripts/teams.json)
"""

import json
import os
import sys
from collections import defaultdict
from datetime import datetime, timedelta, timezone
from pathlib import Path

import openpyxl
import requests
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------
YOUTRACK_TOKEN = os.environ.get("YOUTRACK_TOKEN")
YOUTRACK_URL = os.environ.get("YOUTRACK_URL", "").rstrip("/")
REPORTS_DIR = Path(__file__).resolve().parent.parent / "reports"

GITLAB_TOKEN = os.environ.get("GITLAB_TOKEN")
GITLAB_URL = os.environ.get("GITLAB_URL", "").rstrip("/")

# Load teams config from file or env var
_TEAMS_CONFIG_PATH = os.environ.get("TEAMS_CONFIG", str(Path(__file__).resolve().parent / "teams.json"))
if os.path.isfile(_TEAMS_CONFIG_PATH):
    TEAMS = json.loads(Path(_TEAMS_CONFIG_PATH).read_text())
else:
    print(f"ERROR: Teams config not found at {_TEAMS_CONFIG_PATH}", file=sys.stderr)
    print("Copy scripts/teams.example.json to scripts/teams.json and customize it.", file=sys.stderr)
    sys.exit(1)

if not YOUTRACK_TOKEN:
    print("ERROR: YOUTRACK_TOKEN environment variable is not set", file=sys.stderr)
    sys.exit(1)

if not YOUTRACK_URL:
    print("ERROR: YOUTRACK_URL environment variable is not set", file=sys.stderr)
    sys.exit(1)

API = f"{YOUTRACK_URL}/api"
ISSUE_URL = f"{YOUTRACK_URL}/issue"
HEADERS = {"Authorization": f"Bearer {YOUTRACK_TOKEN}", "Accept": "application/json"}
ISSUE_FIELDS = "idReadable,summary,created,updated,resolved,state(name),assignee(name),customFields(name,value(name,minutes,presentation))"

GL_API = f"{GITLAB_URL}/api/v4" if GITLAB_URL else ""
GL_HEADERS = {"PRIVATE-TOKEN": GITLAB_TOKEN} if GITLAB_TOKEN else {}

# ---------------------------------------------------------------------------
# YouTrack helpers
# ---------------------------------------------------------------------------

def yt_get(endpoint: str, params: dict | None = None) -> list | dict:
    r = requests.get(f"{API}/{endpoint}", headers=HEADERS, params=params or {}, timeout=30)
    r.raise_for_status()
    return r.json()


def fetch_issues(query: str, top: int = 200) -> list[dict]:
    return yt_get("issues", {"query": query, "fields": ISSUE_FIELDS, "$top": top})


def get_field(issue: dict, field_name: str):
    for cf in issue.get("customFields", []):
        if cf["name"] == field_name:
            val = cf.get("value")
            if val is None:
                return None
            if isinstance(val, dict):
                return val.get("name") or val.get("presentation") or val.get("minutes")
            if isinstance(val, list):
                return ", ".join(v.get("name", "") for v in val)
            return val
    return None


def ts_to_date(ts) -> str | None:
    if ts is None:
        return None
    return datetime.fromtimestamp(ts / 1000, tz=timezone.utc).strftime("%Y-%m-%d")


def days_since(ts) -> int | None:
    if ts is None:
        return None
    dt = datetime.fromtimestamp(ts / 1000, tz=timezone.utc)
    return (datetime.now(tz=timezone.utc) - dt).days


def parse_issue(issue: dict) -> dict:
    created, updated, resolved = issue.get("created"), issue.get("updated"), issue.get("resolved")
    estimation = get_field(issue, "Estimation")
    iid = issue["idReadable"]
    state_obj = issue.get("state")
    state = (state_obj.get("name") if isinstance(state_obj, dict) else None) or get_field(issue, "State")
    assignee_obj = issue.get("assignee")
    assignee = (assignee_obj.get("name") if isinstance(assignee_obj, dict) else None) or get_field(issue, "Assignee")
    return {
        "id": iid,
        "url": f"{ISSUE_URL}/{iid}",
        "summary": issue["summary"],
        "state": state,
        "priority": get_field(issue, "Priority") or "None",
        "type": get_field(issue, "Type") or "None",
        "assignee": assignee or "Unassigned",
        "product": get_field(issue, "Product") or "—",
        "estimation": estimation or "NONE",
        "created": ts_to_date(created),
        "updated": ts_to_date(updated),
        "resolved": ts_to_date(resolved),
        "days_since_created": days_since(created),
        "days_since_updated": days_since(updated),
        "is_unestimated": not estimation,
        "is_long_unfinished": (days_since(created) or 0) > 30,
        "is_stale": (days_since(updated) or 0) > 30,
        "is_stuck": state in ("In Progress", "In Review", "Ready for Test") and (days_since(updated) or 0) > 7,
    }


# ---------------------------------------------------------------------------
# GitLab helpers
# ---------------------------------------------------------------------------

def _resolve_gitlab_projects(paths: list[str]) -> list[dict]:
    """Resolve a list of GitLab group/project paths to project objects."""
    if not GL_API:
        return []
    projects = []
    seen_ids = set()

    for path in paths:
        encoded = path.replace("/", "%2F")

        r = requests.get(f"{GL_API}/projects/{encoded}", headers=GL_HEADERS, timeout=30)
        if r.status_code == 200:
            proj = r.json()
            if proj["id"] not in seen_ids:
                projects.append(proj)
                seen_ids.add(proj["id"])
            continue

        page = 1
        while True:
            r = requests.get(f"{GL_API}/groups/{encoded}/projects",
                             headers=GL_HEADERS,
                             params={"per_page": 100, "include_subgroups": "true", "page": page},
                             timeout=30)
            if r.status_code != 200:
                print(f"  WARNING: GitLab path '{path}' not found as project or group")
                break
            batch = r.json()
            if not batch:
                break
            for proj in batch:
                if proj["id"] not in seen_ids:
                    projects.append(proj)
                    seen_ids.add(proj["id"])
            if len(batch) < 100:
                break
            page += 1

    return projects


def fetch_team_gitlab_commits(gitlab_paths: list[str]) -> dict:
    """Fetch daily commit stats from specific GitLab projects/groups."""
    if not GITLAB_TOKEN or not gitlab_paths or not GL_API:
        return {"authors": [], "total_commits": 0}

    projects = _resolve_gitlab_projects(gitlab_paths)
    if not projects:
        return {"authors": [], "total_commits": 0}

    now = datetime.now(tz=timezone.utc)
    since = (now - timedelta(hours=24)).strftime("%Y-%m-%dT%H:%M:%SZ")
    until = now.strftime("%Y-%m-%dT%H:%M:%SZ")

    author_stats = defaultdict(lambda: {"commits": 0, "projects": set(), "last_message": ""})

    for proj in projects:
        pid = proj["id"]
        pname = proj["path_with_namespace"].split("/")[-1]
        page = 1
        while True:
            r = requests.get(f"{GL_API}/projects/{pid}/repository/commits",
                             headers=GL_HEADERS,
                             params={"since": since, "until": until, "per_page": 100, "page": page},
                             timeout=30)
            if r.status_code != 200:
                break
            commits = r.json()
            if not commits or isinstance(commits, dict):
                break
            for c in commits:
                author = c.get("author_name", "Unknown")
                author_stats[author]["commits"] += 1
                author_stats[author]["projects"].add(pname)
                author_stats[author]["last_message"] = c.get("title", "")
            if len(commits) < 100:
                break
            page += 1

    authors = sorted(
        [{"name": name, **data, "projects": sorted(data["projects"])} for name, data in author_stats.items()],
        key=lambda x: -x["commits"]
    )
    total = sum(a["commits"] for a in authors)
    return {"authors": authors, "total_commits": total}


def fetch_dev_activity(username: str, display_name: str) -> dict:
    """Fetch daily activity for a specific GitLab user."""
    if not GITLAB_TOKEN or not GL_API:
        return {}

    r = requests.get(f"{GL_API}/users", headers=GL_HEADERS,
                     params={"username": username}, timeout=30)
    users = r.json()
    if not users:
        print(f"  WARNING: GitLab user '{username}' not found")
        return {}
    user_id = users[0]["id"]

    after_date = (datetime.now(tz=timezone.utc) - timedelta(days=2)).strftime("%Y-%m-%d")
    cutoff = datetime.now(tz=timezone.utc) - timedelta(hours=24)

    all_events = []
    page = 1
    while True:
        r = requests.get(f"{GL_API}/users/{user_id}/events", headers=GL_HEADERS,
                         params={"after": after_date, "per_page": 100, "page": page}, timeout=30)
        if r.status_code != 200:
            break
        batch = r.json()
        if not batch:
            break
        all_events.extend(batch)
        if len(batch) < 100:
            break
        page += 1

    filtered = []
    for e in all_events:
        created = e.get("created_at", "")
        if created:
            evt_time = datetime.fromisoformat(created.replace("Z", "+00:00"))
            if evt_time >= cutoff:
                filtered.append(e)
    all_events = filtered

    activity = defaultdict(int)
    commits, mr_opened, mr_merged, mr_approved = 0, 0, 0, 0
    projects = set()
    for e in all_events:
        action = e.get("action_name", "unknown")
        activity[action] += 1
        if action in ("pushed to", "pushed new"):
            pd = e.get("push_data", {})
            commits += pd.get("commit_count", 0)
        if e.get("target_type") == "MergeRequest":
            if action == "opened":
                mr_opened += 1
            elif action == "accepted":
                mr_merged += 1
            elif action == "approved":
                mr_approved += 1
        pid = e.get("project_id")
        if pid:
            projects.add(pid)

    return {
        "username": username,
        "display_name": display_name,
        "user_id": user_id,
        "last_24h": {
            "total_events": len(all_events),
            "commits": commits,
            "projects_count": len(projects),
            "mr_opened": mr_opened,
            "mr_merged": mr_merged,
            "mr_approved": mr_approved,
            "activity": dict(activity),
        },
    }


# ---------------------------------------------------------------------------
# Excel styling
# ---------------------------------------------------------------------------
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
LINK_FONT = Font(color="0563C1", underline="single")
RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
ORANGE = PatternFill(start_color="FFE0B2", end_color="FFE0B2", fill_type="solid")
YELLOW = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
GREEN = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
BLUE = PatternFill(start_color="BBDEFB", end_color="BBDEFB", fill_type="solid")
LIGHT_BLUE = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
PURPLE = PatternFill(start_color="E1BEE7", end_color="E1BEE7", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin", color="D0D0D0"), right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"), bottom=Side(style="thin", color="D0D0D0"),
)


def write_sheet(ws, data, title, headers, row_fn, color_fn=None, widths=None):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws.cell(row=1, column=1, value=title).font = Font(bold=True, size=14, color="2F5496")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    ts = datetime.now().strftime("%Y-%m-%d %H:%M")
    ws.cell(row=2, column=1, value=f"Total: {len(data)} | Generated: {ts}").font = Font(
        italic=True, color="666666", size=10
    )
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=col, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center")
        c.border = THIN_BORDER
    for ri, item in enumerate(data, 5):
        for col, val in enumerate(row_fn(item), 1):
            c = ws.cell(row=ri, column=col, value=val)
            c.border = THIN_BORDER
            c.alignment = Alignment(vertical="center", wrap_text=(col == 2))
        ic = ws.cell(row=ri, column=1)
        ic.hyperlink = item.get("url", f"{ISSUE_URL}/{item.get('id', '')}")
        ic.font = LINK_FONT
        if color_fn:
            fill = color_fn(item)
            if fill:
                for col in range(1, len(headers) + 1):
                    ws.cell(row=ri, column=col).fill = fill
                ws.cell(row=ri, column=1).font = LINK_FONT
    ws.freeze_panes = "A5"
    if data:
        ws.auto_filter.ref = f"A4:{get_column_letter(len(headers))}{4 + len(data)}"
    if widths:
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w


# ---------------------------------------------------------------------------
# Snapshots — per-team
# ---------------------------------------------------------------------------

def _team_slug(team_name: str) -> str:
    return team_name.lower().replace(" ", "_")


def load_team_snapshot(team_name: str) -> dict[str, dict]:
    slug = _team_slug(team_name)
    snap = REPORTS_DIR / f"latest_snapshot_{slug}.json"
    if snap.exists():
        return json.loads(snap.read_text())
    return {}


def save_team_snapshot(team_name: str, parsed: dict[str, dict]):
    slug = _team_slug(team_name)
    snap = REPORTS_DIR / f"latest_snapshot_{slug}.json"
    snap.write_text(json.dumps(parsed, ensure_ascii=False, indent=2))


# ---------------------------------------------------------------------------
# Process a single team
# ---------------------------------------------------------------------------

def process_team(team: dict) -> dict:
    """Fetch and analyze issues + GitLab commits for one team."""
    name = team["name"]
    yt_queries = team["yt_query"]
    if isinstance(yt_queries, str):
        yt_queries = [yt_queries]

    print(f"\n{'='*60}")
    print(f"Processing team: {name}")
    print(f"{'='*60}")

    # YT unresolved
    raw_unresolved = []
    seen_ids = set()
    for q in yt_queries:
        print(f"  Fetching unresolved issues for: {q}")
        issues = fetch_issues(f"{q} #Unresolved")
        for i in issues:
            if i["idReadable"] not in seen_ids:
                raw_unresolved.append(i)
                seen_ids.add(i["idReadable"])
    print(f"    Total: {len(raw_unresolved)} unresolved issues")

    # YT recently resolved
    today_str = datetime.now().strftime("%Y-%m-%d")
    week_ago = (datetime.now() - timedelta(days=7)).strftime("%Y-%m-%d")
    raw_resolved = []
    seen_ids = set()
    print(f"  Fetching recently resolved issues (last 7 days)")
    for q in yt_queries:
        issues = fetch_issues(f"{q} resolved date: {week_ago} .. {today_str}")
        for i in issues:
            if i["idReadable"] not in seen_ids:
                raw_resolved.append(i)
                seen_ids.add(i["idReadable"])
    print(f"    Total: {len(raw_resolved)} recently resolved issues")

    # GitLab commits
    gitlab_paths = team.get("gitlab_paths", [])
    gl_stats = {"authors": [], "total_commits": 0}
    if gitlab_paths:
        print(f"  Fetching GitLab commits for: {', '.join(gitlab_paths)}")
        gl_stats = fetch_team_gitlab_commits(gitlab_paths)
        print(f"    {gl_stats['total_commits']} commits by {len(gl_stats['authors'])} authors")

    # Parse & diff
    new_parsed = {i["idReadable"]: parse_issue(i) for i in raw_unresolved}
    resolved_parsed = {i["idReadable"]: parse_issue(i) for i in raw_resolved}
    old_parsed = load_team_snapshot(name)

    old_ids = set(old_parsed.keys())
    new_ids = set(new_parsed.keys())
    closed_ids = old_ids - new_ids
    new_issue_ids = new_ids - old_ids

    state_changes, updated_only = [], []
    for iid in old_ids & new_ids:
        os_ = old_parsed[iid]["state"]
        ns_ = new_parsed[iid]["state"]
        ou_ = old_parsed[iid]["updated"]
        nu_ = new_parsed[iid]["updated"]
        if os_ != ns_:
            state_changes.append({"old_state": os_, **new_parsed[iid]})
        elif ou_ != nu_:
            updated_only.append({"old_updated": ou_, "new_updated": nu_, **new_parsed[iid]})

    has_previous = bool(old_parsed)

    # Counts
    nu_cnt = len([p for p in new_parsed.values() if p["is_unestimated"]])
    ou_cnt = len([p for p in old_parsed.values() if p.get("is_unestimated")]) if has_previous else 0
    nl_cnt = len([p for p in new_parsed.values() if p["is_long_unfinished"]])
    ol_cnt = len([p for p in old_parsed.values() if p.get("is_long_unfinished")]) if has_previous else 0
    ns_cnt = len([p for p in new_parsed.values() if p["is_stale"]])
    os_cnt = len([p for p in old_parsed.values() if p.get("is_stale")]) if has_previous else 0

    # Tracked developer stats
    tracked_devs = team.get("tracked_devs", {})
    dev_stats = {}
    for username, display_name in tracked_devs.items():
        print(f"  Fetching activity for: {display_name} (@{username})")
        dev_stats[username] = fetch_dev_activity(username, display_name)
        ds = dev_stats[username]
        if ds:
            t = ds["last_24h"]
            print(f"    Last 24h: {t['commits']} commits, {t['mr_opened']} MRs opened, "
                  f"{t['mr_merged']} merged, {t['mr_approved']} approved")

    save_team_snapshot(name, new_parsed)
    print(f"  Snapshot saved ({len(new_parsed)} issues)")

    return {
        "team": team,
        "new_parsed": new_parsed,
        "resolved_parsed": resolved_parsed,
        "old_parsed": old_parsed,
        "has_previous": has_previous,
        "old_ids": old_ids,
        "new_ids": new_ids,
        "closed_ids": closed_ids,
        "new_issue_ids": new_issue_ids,
        "state_changes": state_changes,
        "updated_only": updated_only,
        "gl_stats": gl_stats,
        "dev_stats": dev_stats,
        "counts": {
            "nu": nu_cnt, "ou": ou_cnt,
            "nl": nl_cnt, "ol": ol_cnt,
            "ns": ns_cnt, "os": os_cnt,
        },
    }


# ---------------------------------------------------------------------------
# Excel + HTML builders (imported from original, no hardcoded data)
# ---------------------------------------------------------------------------

def _sheet_name(team_name: str, suffix: str) -> str:
    prefix = team_name[:12]
    name = f"{prefix} - {suffix}"
    return name[:31]


def _resolve_boards(team: dict) -> list[dict]:
    boards = team.get("yt_boards", [])
    if boards:
        return boards
    urls = team.get("yt_board_urls", [])
    if not urls:
        single = team.get("yt_board_url", "")
        urls = [single] if single else []
    return [{"name": f"Board {i+1}", "url": u} for i, u in enumerate(urls)]


def build_excel(team_results: list[dict]) -> openpyxl.Workbook:
    """Build multi-team Excel workbook. See original for full implementation."""
    wb = openpyxl.Workbook()
    first_sheet = True

    for tr in team_results:
        team = tr["team"]
        tname = team["name"]
        new_parsed = tr["new_parsed"]
        new_ids = tr["new_ids"]
        c = tr["counts"]
        ts_str = datetime.now().strftime("%Y-%m-%d %H:%M")

        # All Unresolved sheet
        if first_sheet:
            ws = wb.active
            ws.title = _sheet_name(tname, "All Unresolved")
            first_sheet = False
        else:
            ws = wb.create_sheet(_sheet_name(tname, "All Unresolved"))

        all_current = sorted(new_parsed.values(), key=lambda x: -(x["days_since_created"] or 0))
        headers = ["Issue ID", "Summary", "Product", "State", "Priority", "Type",
                    "Assignee", "Estimation", "Created", "Updated", "Age (days)",
                    "Days Since Update", "Flags"]

        def row_fn(i):
            fl = []
            if i["is_unestimated"]:
                fl.append("NO ESTIMATE")
            if i.get("is_stuck"):
                fl.append("STUCK")
            if i["is_stale"]:
                fl.append("STALE")
            if i["days_since_created"] and i["days_since_created"] > 200:
                fl.append("ANCIENT")
            elif i["is_long_unfinished"]:
                fl.append("LONG OPEN")
            return [i["id"], i["summary"], i["product"], i["state"], i["priority"],
                    i["type"], i["assignee"], i["estimation"], i["created"], i["updated"],
                    i["days_since_created"], i["days_since_updated"], ", ".join(fl)]

        def color_fn(i):
            if i["days_since_created"] and i["days_since_created"] > 200:
                return RED
            if i.get("is_stuck"):
                return RED
            if i["is_stale"]:
                return ORANGE
            if i["is_long_unfinished"]:
                return YELLOW
            return None

        write_sheet(ws, all_current, f"{tname} — All Unresolved Issues", headers,
                    row_fn, color_fn,
                    [12, 60, 20, 16, 10, 14, 35, 12, 12, 12, 12, 16, 25])

        # Summary sheet
        ws_sum = wb.create_sheet(_sheet_name(tname, "Summary"))
        ws_sum.merge_cells("A1:D1")
        ws_sum.cell(row=1, column=1, value=f"{tname} — Issue Health Summary").font = Font(
            bold=True, size=14, color="2F5496"
        )
        ws_sum.cell(row=2, column=1, value=f"Generated: {ts_str}").font = Font(italic=True, color="666666")
        for col, h in enumerate(["Metric", "Count", "% of Total", "Severity"], 1):
            cell = ws_sum.cell(row=4, column=col, value=h)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.border = THIN_BORDER

        total = len(new_ids) or 1
        blocked = len([p for p in new_parsed.values() if p["state"] == "Blocked"])
        unassigned = len([p for p in new_parsed.values() if p["assignee"] == "Unassigned"])
        ancient = len([p for p in new_parsed.values() if (p["days_since_created"] or 0) > 200])
        stuck = len([p for p in new_parsed.values() if p.get("is_stuck")])

        sd = [
            ("Total Unresolved", len(new_ids), "100%", "—"),
            ("Unestimated", c["nu"], f"{c['nu'] * 100 // total}%", "CRITICAL"),
            ("Stuck In Progress (>7d)", stuck, f"{stuck * 100 // total}%", "CRITICAL"),
            ("Long Unfinished (>30d)", c["nl"], f"{c['nl'] * 100 // total}%", "HIGH"),
            ("Ancient (>200d)", ancient, f"{ancient * 100 // total}%", "CRITICAL"),
            ("Stale (no update >30d)", c["ns"], f"{c['ns'] * 100 // total}%", "HIGH"),
            ("Blocked", blocked, f"{blocked * 100 // total}%", "MEDIUM"),
            ("Unassigned", unassigned, f"{unassigned * 100 // total}%", "MEDIUM"),
        ]
        sev_colors = {"CRITICAL": RED, "HIGH": ORANGE, "MEDIUM": YELLOW, "—": None}
        for ri, (m, cnt, pct, sev) in enumerate(sd, 5):
            for col, v in enumerate([m, cnt, pct, sev], 1):
                ws_sum.cell(row=ri, column=col, value=v).border = THIN_BORDER
            if sev_colors.get(sev):
                for cc in range(1, 5):
                    ws_sum.cell(row=ri, column=cc).fill = sev_colors[sev]
        for i, w in enumerate([35, 12, 12, 14], 1):
            ws_sum.column_dimensions[get_column_letter(i)].width = w

    return wb


def build_html_email(team_results: list[dict]) -> str:
    """Build HTML email body for all teams."""
    today = datetime.now().strftime("%Y-%m-%d")

    def issue_link(iid):
        return f'<a href="{ISSUE_URL}/{iid}">{iid}</a>'

    def delta_badge(val):
        if isinstance(val, int):
            if val < 0:
                return f'<span style="color:#2e7d32;font-weight:bold">{val}</span>'
            if val > 0:
                return f'<span style="color:#c62828;font-weight:bold">+{val}</span>'
        return str(val)

    TEAM_COLORS = ["#2F5496", "#1B5E20", "#BF360C", "#4A148C", "#006064",
                   "#E65100", "#1A237E", "#880E4F", "#33691E", "#311B92"]

    html_parts = [
        '<!DOCTYPE html><html><head><meta charset="utf-8"></head>',
        '<body style="font-family:Arial,sans-serif;font-size:14px;color:#333;max-width:900px;margin:0 auto">',
        f'<h1 style="color:#2F5496;border-bottom:2px solid #2F5496;padding-bottom:8px">'
        f'Daily Report — {today}</h1>',
    ]

    # Table of contents
    html_parts.append('<div style="background:#f5f5f5;padding:12px 16px;border-radius:6px;margin-bottom:20px">')
    html_parts.append('<b>Teams:</b> ')
    for i, tr in enumerate(team_results):
        tname = tr["team"]["name"]
        color = TEAM_COLORS[i % len(TEAM_COLORS)]
        sep = " | " if i > 0 else ""
        boards = _resolve_boards(tr["team"])
        if len(boards) == 1:
            boards_html = f' (<a href="{boards[0]["url"]}" style="color:#888;font-size:12px">board</a>)'
        elif boards:
            links = " ".join(
                f'<a href="{b["url"]}" style="color:#888;font-size:12px">[{b["name"]}]</a>' for b in boards
            )
            boards_html = f' ({links})'
        else:
            boards_html = ""
        html_parts.append(
            f'{sep}<a href="#{tname.lower()}" style="color:{color};font-weight:bold">{tname}</a>'
            f'{boards_html}'
        )
    html_parts.append('</div>')

    # Per-team sections
    for idx, tr in enumerate(team_results):
        team = tr["team"]
        tname = team["name"]
        color = TEAM_COLORS[idx % len(TEAM_COLORS)]
        new_parsed = tr["new_parsed"]
        has_previous = tr["has_previous"]
        old_ids, new_ids = tr["old_ids"], tr["new_ids"]
        closed_ids, new_issue_ids = tr["closed_ids"], tr["new_issue_ids"]
        state_changes = tr["state_changes"]
        c = tr["counts"]

        html_parts.append(
            f'<div id="{tname.lower()}" style="margin-top:30px">'
            f'<h2 style="color:{color};border-bottom:2px solid {color};padding-bottom:6px">{tname}</h2>'
        )

        # Overview
        html_parts.append('<h3>Overview</h3>')
        html_parts.append('<table style="border-collapse:collapse;width:100%">')
        html_parts.append(
            f'<tr style="background:{color};color:#fff">'
            '<th style="padding:8px;text-align:left">Metric</th>'
            '<th style="padding:8px">Previous</th><th style="padding:8px">Current</th>'
            '<th style="padding:8px">Delta</th></tr>'
        )

        if has_previous:
            rows = [
                ("Total Unresolved", len(old_ids), len(new_ids), len(new_ids) - len(old_ids)),
                ("Unestimated", c["ou"], c["nu"], c["nu"] - c["ou"]),
                ("Long Unfinished (>30d)", c["ol"], c["nl"], c["nl"] - c["ol"]),
                ("Stale (no update >30d)", c["os"], c["ns"], c["ns"] - c["os"]),
            ]
        else:
            rows = [
                ("Total Unresolved", "—", len(new_ids), "—"),
                ("Unestimated", "—", c["nu"], "—"),
                ("Long Unfinished (>30d)", "—", c["nl"], "—"),
                ("Stale (no update >30d)", "—", c["ns"], "—"),
            ]

        for i, (m, prev, curr, delta) in enumerate(rows):
            bg = "#f5f5f5" if i % 2 else "#fff"
            html_parts.append(
                f'<tr style="background:{bg}">'
                f'<td style="padding:6px 8px;border-bottom:1px solid #e0e0e0">{m}</td>'
                f'<td style="padding:6px 8px;text-align:center;border-bottom:1px solid #e0e0e0">{prev}</td>'
                f'<td style="padding:6px 8px;text-align:center;border-bottom:1px solid #e0e0e0"><b>{curr}</b></td>'
                f'<td style="padding:6px 8px;text-align:center;border-bottom:1px solid #e0e0e0">{delta_badge(delta)}</td>'
                f'</tr>'
            )
        html_parts.append('</table>')

        if has_previous:
            if closed_ids:
                html_parts.append(f'<h3 style="color:#2e7d32">Resolved/Closed ({len(closed_ids)})</h3><ul>')
                for iid in sorted(closed_ids):
                    r = tr["resolved_parsed"].get(iid, tr["old_parsed"].get(iid, {}))
                    html_parts.append(f'<li>{issue_link(iid)}: {r.get("summary", "")}</li>')
                html_parts.append('</ul>')

            if state_changes:
                html_parts.append(f'<h3 style="color:#1565c0">State Changes ({len(state_changes)})</h3><ul>')
                for s in sorted(state_changes, key=lambda x: x["id"]):
                    html_parts.append(
                        f'<li>{issue_link(s["id"])}: {s["summary"]} '
                        f'<span style="color:#888">({s["old_state"]} &rarr; {s["state"]})</span></li>'
                    )
                html_parts.append('</ul>')

            if new_issue_ids:
                html_parts.append(f'<h3 style="color:#0277bd">New Issues ({len(new_issue_ids)})</h3><ul>')
                for iid in sorted(new_issue_ids):
                    n = new_parsed[iid]
                    html_parts.append(
                        f'<li>{issue_link(iid)}: {n["summary"]} '
                        f'<span style="color:#888">[{n["state"]}] {n["priority"]}</span></li>'
                    )
                html_parts.append('</ul>')
        else:
            html_parts.append('<p><i>First run — no comparison with previous report.</i></p>')

        html_parts.append('</div>')

    html_parts.append(
        '<hr style="margin-top:20px;border:none;border-top:1px solid #e0e0e0">'
        '<p style="color:#999;font-size:12px">Generated automatically by yt-mcp report script.</p>'
        '</body></html>'
    )

    return "\n".join(html_parts)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    REPORTS_DIR.mkdir(parents=True, exist_ok=True)

    team_results = []
    for team in TEAMS:
        tr = process_team(team)
        team_results.append(tr)

    # Build Excel
    wb = build_excel(team_results)
    today = datetime.now().strftime("%Y-%m-%d")
    out_path = REPORTS_DIR / f"youtrack_issues_{today}.xlsx"
    wb.save(out_path)
    print(f"\nReport saved: {out_path}")
    latest_path = REPORTS_DIR / "youtrack_issues_latest.xlsx"
    wb.save(latest_path)

    # Build HTML email
    email_html = build_html_email(team_results)
    email_path = REPORTS_DIR / "email_body.html"
    email_path.write_text(email_html)
    print(f"Email body saved: {email_path}")

    # Print summary
    print(f"\n{'='*60}")
    print(f"REPORT SUMMARY")
    print(f"{'='*60}")
    for tr in team_results:
        tname = tr["team"]["name"]
        print(f"\n  [{tname}]")
        print(f"    Unresolved: {len(tr['new_ids'])}")
        if tr["has_previous"]:
            print(f"    Previous:   {len(tr['old_ids'])}")
            print(f"    Resolved:   {len(tr['closed_ids'])}")
            print(f"    New:        {len(tr['new_issue_ids'])}")
            print(f"    Changed:    {len(tr['state_changes'])} state + {len(tr['updated_only'])} updated")
        else:
            print("    (First run — no comparison available)")
        c = tr["counts"]
        print(f"    Unestimated: {c['nu']}")
        print(f"    Long (>30d): {c['nl']}")
        print(f"    Stale:       {c['ns']}")
        if tr["gl_stats"]["total_commits"]:
            print(f"    GitLab:      {tr['gl_stats']['total_commits']} commits")


if __name__ == "__main__":
    main()
